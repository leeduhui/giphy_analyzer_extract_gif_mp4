import pandas as pd
from tqdm import tqdm
import os
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

# 첫 번째 CSV 파일 로드
df_views = pd.read_csv("GIPHY_view_count_eq4all_gif_sticker_2013-02-01_2024-10-26.csv")

# content_id 컬럼 추출 및 결측치 제거
content_ids = df_views['content_id'].dropna().astype(str)

# 기본 URL 설정
#base_url = "https://i.giphy.com/media/v1.Y2lkPTc5MGI3NjExc..."
base_url = "https://i.giphy.com/media/v1.Y2lkPTc5MGI3NjExcHRmNXo3MWZ2ZmNmYnI0aWVqZGc0NXlyenRxeTN0bjA5OHh0NmR0biZlcD12MV9pbnRlcm5hbF9naWZfYnlfaWQmY3Q9Zw/"

# 두 번째 CSV 파일 로드
df_engagement = pd.read_csv("GIPHY_engagement_rate_eq4all_gif_sticker_2013-02-01_2024-10-26.csv")

# content_id와 term의 매핑 생성
term_dict = dict(zip(df_engagement['content_id'].astype(str), df_engagement['term']))

# 결과를 저장할 리스트 초기화
data = []

# 다운로드 폴더 생성
gif_download_folder = 'download_gif'
png_download_folder = 'download_png'  # PNG 이미지 저장 폴더

os.makedirs(gif_download_folder, exist_ok=True)
os.makedirs(png_download_folder, exist_ok=True)

# 각 content_id에 대해 URL 생성 및 파일 다운로드
for content_id in tqdm(content_ids, desc="URL 생성 및 파일 다운로드 중"):
    try:
        # 문자열로 변환하여 안전하게 처리
        content_id_str = content_id.strip()

        mid_url = f"{base_url}{content_id_str}"

        # URL 생성
        full_gif_url = os.path.join(mid_url, "giphy.gif")
        full_mp4_url = os.path.join(mid_url, "giphy.mp4")

        # 해당 content_id의 term 가져오기
        term = term_dict.get(content_id_str, "No Term")

        # 데이터 리스트에 추가
        data.append({
            'term': term,
            'gif_image_path': '',  # GIF 이미지의 PNG 변환 파일 경로
            'content_id': content_id_str,
            'link_gif_url': '',
            'link_mp4_url': '',
            'full_gif_url': full_gif_url,
            'full_mp4_url': full_mp4_url
        })

        # 파일 경로 설정
        gif_file_path = os.path.join(gif_download_folder, f"{content_id_str}.gif")
        png_file_path = os.path.join(png_download_folder, f"{content_id_str}.png")

        # GIF 파일 다운로드 (이미 존재하면 스킵)
        if not os.path.exists(gif_file_path):
            gif_response = requests.get(full_gif_url, stream=True)
            if gif_response.status_code == 200:
                with open(gif_file_path, 'wb') as f:
                    for chunk in gif_response.iter_content(chunk_size=1024):
                        if chunk:
                            f.write(chunk)
            else:
                print(f"GIF 파일을 다운로드할 수 없습니다: {full_gif_url}")
        else:
            print(f"{gif_file_path} 이미 존재합니다. 다운로드를 건너뜁니다.")

        # GIF를 PNG로 변환하여 저장 (이미 존재하면 스킵)
        if not os.path.exists(png_file_path):
            try:
                with PILImage.open(gif_file_path) as im:
                    # 중간 프레임 계산
                    middle_frame = im.n_frames // 2
                    print(f"content_id '{content_id_str}'의 middle_frame: {middle_frame}")
                    im.seek(middle_frame)  # 중간 프레임으로 이동

                    # 이미지의 크기 가져오기
                    original_width, original_height = im.size
                    new_size = (original_width // 2, original_height // 2)  # 크기를 절반으로 줄임

                    # 이미지 크기 조정
                    im_resized = im.resize(new_size, PILImage.ANTIALIAS)

                    # PNG로 저장
                    im_resized.save(png_file_path)
            except Exception as e:
                print(f"GIF를 PNG로 변환하는 중 오류 발생: {e}")
                png_file_path = ''  # 변환 실패 시 빈 문자열로 설정
        else:
            print(f"{png_file_path} 이미 존재합니다. 변환을 건너뜁니다.")

        # gif_image_path 업데이트
        data[-1]['gif_image_path'] = png_file_path

    except Exception as e:
        print(f"content_id '{content_id_str}'에서 오류 발생: {e}")

# 데이터프레임 생성
result_df = pd.DataFrame(data)

# 하이퍼링크 생성 함수
def make_hyperlink(url):
    return f'=HYPERLINK("{url}", "링크")'

# full_gif_url과 full_mp4_url 열을 하이퍼링크로 변환하여 link_gif_url과 link_mp4_url에 저장
result_df['link_gif_url'] = result_df['full_gif_url'].apply(make_hyperlink)
result_df['link_mp4_url'] = result_df['full_mp4_url'].apply(make_hyperlink)

# 열 순서 재정렬
result_df = result_df[['term', 'gif_image_path', 'content_id', 'link_gif_url', 'link_mp4_url', 'full_gif_url', 'full_mp4_url']]

# Excel 파일로 저장 (이미지 포함)
excel_file = "result_urls.xlsx"

# 새로운 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "GIF Data"

# 열 제목 추가
ws.append(['term', 'gif', 'content_id', 'link_gif_url', 'link_mp4_url', 'full_gif_url', 'full_mp4_url'])

# 각 행에 데이터 추가
for index, row in result_df.iterrows():
    term = row['term']
    content_id = row['content_id']
    gif_image_path = row['gif_image_path']
    link_gif_url = row['link_gif_url']
    link_mp4_url = row['link_mp4_url']
    full_gif_url = row['full_gif_url']
    full_mp4_url = row['full_mp4_url']

    # 현재 행 번호 계산 (헤더가 있으므로 +2)
    current_row = index + 2

    # 데이터 추가 (이미지와 링크 셀은 나중에 처리)
    ws.cell(row=current_row, column=1, value=term)        # term
    ws.cell(row=current_row, column=3, value=content_id)  # content_id

    # 이미지 삽입 및 셀 크기 조정
    if gif_image_path and os.path.exists(gif_image_path):
        img = Image(gif_image_path)

        # 이미지의 실제 크기 (픽셀 단위)
        img_width, img_height = img.width, img.height

        # 이미지 삽입 (현재 행의 B열)
        ws.add_image(img, f'B{current_row}')

        # 열 너비 및 행 높이 조정
        # 픽셀을 Excel의 열 너비와 행 높이 단위로 변환
        def pixels_to_width(pixels):
            return pixels * 0.14  # 경험적인 값

        def pixels_to_height(pixels):
            return pixels * 0.75  # 경험적인 값

        column_letter = 'B'
        column_dim = ws.column_dimensions[column_letter]
        row_dim = ws.row_dimensions[current_row]

        # 열 너비 설정 (이미 설정된 너비보다 작을 때만 업데이트)
        current_width = column_dim.width or 0
        new_width = pixels_to_width(img_width)
        if new_width < current_width or current_width == 0:
            column_dim.width = new_width

        # 행 높이 설정
        row_dim.height = pixels_to_height(img_height)
    else:
        print(f"이미지 파일이 존재하지 않습니다: {gif_image_path}")

    # 링크_gif_url 셀 설정
    gif_url_cell = ws.cell(row=current_row, column=4)  # D열
    gif_url_cell.value = "링크"
    gif_url_cell.hyperlink = full_gif_url
    gif_url_cell.style = "Hyperlink"

    # 링크_mp4_url 셀 설정
    mp4_url_cell = ws.cell(row=current_row, column=5)  # E열
    mp4_url_cell.value = "링크"
    mp4_url_cell.hyperlink = full_mp4_url
    mp4_url_cell.style = "Hyperlink"

    # full_gif_url 및 full_mp4_url 셀 설정
    ws.cell(row=current_row, column=6, value=full_gif_url)  # F열
    ws.cell(row=current_row, column=7, value=full_mp4_url)  # G열

# 워크북 저장
wb.save(excel_file)

print("URL 생성, 파일 다운로드 및 Excel 파일 저장이 완료되었습니다.")

